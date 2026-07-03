"""
audit.py — Two-stage adversarial auditor for supplier verification.

The auditor receives Stage 1's verdicts and the same web-evidence block,
then challenges those verdicts. Import into verify.py via:

    from audit import ai_audit_row, write_audit_headers, write_audit_row, AUDIT_COL_COUNT

Also exposes `audit_bp` — a Flask Blueprint serving the standalone /dev/audit page.
"""

import io
import re
import tempfile
import threading
import time
import uuid

import openpyxl
from flask import Blueprint, Response, jsonify, request
from openpyxl.styles import Alignment, Font, PatternFill

from ai import call_ai, safe_parse_json
from scraper import crawl_urls

audit_bp = Blueprint("audit", __name__)

# ── Fills ─────────────────────────────────────────────────────────────────────

AUDIT_HDR_FILL  = PatternFill("solid", fgColor="4A148C")  # deep purple header
AUDIT_CONF_FILL = PatternFill("solid", fgColor="C8E6C9")  # green  — Confirmed
AUDIT_DISP_FILL = PatternFill("solid", fgColor="FFCDD2")  # red    — Disputed
AUDIT_UNCF_FILL = PatternFill("solid", fgColor="FFF9C4")  # yellow — Unconfirmed
AUDIT_HDR_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

AUDIT_COL_COUNT = 6  # columns written per row by write_audit_row()

_VALID_VERDICTS = {"Confirmed", "Disputed", "Unconfirmed"}
_WRAP = Alignment(wrap_text=True, vertical="top")


# ── Internal helpers (re-implemented here to avoid circular imports) ──────────

def _make_logger(log_fn):
    def _log(msg, is_error=False):
        print(msg)
        if log_fn:
            log_fn(msg, is_error)
    return _log


def _call_ai_json(prompt: str, provider: str, max_tokens: int,
                  required_key: str, log) -> dict | None:
    for attempt in range(1, 4):
        try:
            raw    = call_ai(prompt, provider, max_tokens=max_tokens)
            parsed = safe_parse_json(raw)
            if isinstance(parsed, dict) and required_key in parsed:
                return parsed
            log(f"[llm] Attempt {attempt}: response missing '{required_key}' — retrying", True)
        except Exception as e:
            log(f"[llm] Attempt {attempt} failed: {e}", True)
        if attempt < 3:
            time.sleep(5 * attempt)
    log("[llm] All 3 attempts failed", True)
    return None


# ── Fill helper ───────────────────────────────────────────────────────────────

def audit_fill(val: str) -> PatternFill:
    """Map an audit verdict string to its cell fill colour."""
    return {"Confirmed": AUDIT_CONF_FILL, "Disputed": AUDIT_DISP_FILL}.get(val, AUDIT_UNCF_FILL)


# ── Column metadata ───────────────────────────────────────────────────────────

def audit_col_headers(audit_provider: str) -> list[tuple[str, int]]:
    """Return [(header_name, column_width), ...] for the 6 audit columns."""
    label = audit_provider.capitalize()
    return [
        (f"Audit: Company Exists",           16),
        (f"Audit: Supply Ties Exist",         18),
        (f"Audit: Correct Component",         22),
        (f"Audit Notes [{label}]",            55),
        (f"Audit Verdict [{label}]",          18),
        (f"Audit Verdict Reason [{label}]",   45),
    ]


# ── Sheet helpers ─────────────────────────────────────────────────────────────

def write_audit_headers(ws, start_col: int, audit_provider: str) -> None:
    """Write the purple header row for the 6 audit columns starting at start_col."""
    for i, (hdr, width) in enumerate(audit_col_headers(audit_provider)):
        cell = ws.cell(row=1, column=start_col + i, value=hdr)
        cell.fill = AUDIT_HDR_FILL
        cell.font = AUDIT_HDR_FONT
        ws.column_dimensions[cell.column_letter].width = width


def write_audit_row(ws, row_idx: int, start_col: int, result: dict) -> None:
    """Write the 6 audit result cells for a single data row."""
    # Per-check verdict cells (columns 0–2)
    for k, key in enumerate(["audit_company_exists", "audit_supply_ties", "audit_correct_component"]):
        val = result.get(key, "")
        c   = ws.cell(row=row_idx, column=start_col + k, value=val)
        c.alignment = _WRAP
        if val:
            c.fill = audit_fill(val)

    # Combined notes (column 3)
    parts = []
    if result.get("audit_notes_exists"):
        parts.append(f"[Exists] {result['audit_notes_exists']}")
    if result.get("audit_notes_supply"):
        parts.append(f"[Supply] {result['audit_notes_supply']}")
    if result.get("audit_notes_component"):
        parts.append(f"[Component] {result['audit_notes_component']}")
    c = ws.cell(row=row_idx, column=start_col + 3, value=" | ".join(parts))
    c.alignment = _WRAP

    # Overall verdict (column 4) — Partial maps to yellow (Unconfirmed colour)
    verdict_val = result.get("audit_verdict", "")
    c = ws.cell(row=row_idx, column=start_col + 4, value=verdict_val)
    c.alignment = _WRAP
    if verdict_val:
        c.fill = audit_fill(verdict_val if verdict_val != "Partial" else "Unconfirmed")

    # Verdict reason (column 5)
    c = ws.cell(row=row_idx, column=start_col + 5, value=result.get("audit_verdict_reason", ""))
    c.alignment = _WRAP


# ── Core audit function ───────────────────────────────────────────────────────

def ai_audit_row(
    evidence: str,
    company_name: str,
    supplies_to: str,
    components: str,
    stage1: dict,
    audit_provider: str,
    log_fn=None,
    mode: str = "evidence",
) -> dict:
    """
    Adversarial second-pass: challenge Stage 1 verdicts.

    mode="evidence"  — evidence is raw scraped web content (combined verify+audit flow).
    mode="reasoning" — evidence is Stage 1's documented notes; auditor uses training
                       knowledge to challenge whether the reasoning is sound (standalone audit).

    stage1 keys: company_exists (bool), supply_ties (bool), correct_component (bool),
                 notes_exists, notes_supply, notes_component
    """
    _log = _make_logger(log_fn)

    stage1_summary = (
        f"- Company Exists: {stage1['company_exists']} — \"{stage1['notes_exists']}\"\n"
        f"- Supply Ties Exist: {stage1['supply_ties']} — \"{stage1['notes_supply']}\"\n"
        f"- Correct Component Supplied: {stage1['correct_component']} — \"{stage1['notes_component']}\""
    )

    if mode == "reasoning":
        evidence_section = f"""STAGE 1 DOCUMENTED REASONING:
{evidence[:5000]}

INSTRUCTIONS: Audit each of Stage 1's three verdicts by challenging its reasoning.
Draw on your training knowledge of this company, industry, and supply chain relationships.
Be adversarial — look for weak inferences, unjustified conclusions, or claims that contradict known facts."""
    else:
        evidence_section = f"""EVIDENCE (same web evidence used by Stage 1):
{evidence[:5000]}

INSTRUCTIONS: Audit each of Stage 1's three verdicts. Be adversarial — look for weaknesses, gaps, or overreach."""

    prompt = f"""You are a skeptical supply chain auditor. A previous AI model assessed a supplier and you must challenge its conclusions.

SUPPLIER UNDER REVIEW:
- Company: {company_name}
- Supplies to: {supplies_to}
- Components: {components}

STAGE 1 VERDICT:
{stage1_summary}

{evidence_section}

For each verdict use exactly one of these three values:
- "Confirmed": the reasoning is sound and the conclusion is well-justified
- "Disputed": the reasoning is flawed, the conclusion is unjustified, or it contradicts known facts
- "Unconfirmed": the reasoning is too thin or ambiguous to be certain either way

Return ONLY valid JSON (no markdown):
{{
  "audit_company_exists":    "Confirmed" or "Disputed" or "Unconfirmed",
  "audit_supply_ties":       "Confirmed" or "Disputed" or "Unconfirmed",
  "audit_correct_component": "Confirmed" or "Disputed" or "Unconfirmed",
  "audit_notes_exists":      "2-3 sentences: assess the quality of Stage 1's reasoning and cite any supporting or contradicting knowledge",
  "audit_notes_supply":      "2-3 sentences: assess the quality of Stage 1's reasoning and cite any supporting or contradicting knowledge",
  "audit_notes_component":   "2-3 sentences: assess the quality of Stage 1's reasoning and cite any supporting or contradicting knowledge"
}}"""

    _log(f"[audit] Sending to {audit_provider} for '{company_name}'")
    parsed = _call_ai_json(prompt, audit_provider, 1200, "audit_company_exists", _log)

    def _safe(val):
        return val if val in _VALID_VERDICTS else "Unconfirmed"

    if parsed:
        ace = _safe(parsed.get("audit_company_exists"))
        ast = _safe(parsed.get("audit_supply_ties"))
        acc = _safe(parsed.get("audit_correct_component"))
        vals = {ace, ast, acc}

        if "Disputed" in vals:
            verdict = "Disputed"
        elif vals == {"Confirmed"}:
            verdict = "Confirmed"
        elif vals == {"Unconfirmed"}:
            verdict = "Unconfirmed"
        else:
            verdict = "Partial"   # mixed: some Confirmed, some Unconfirmed

        checks = [("Company Exists", ace), ("Supply Ties Exist", ast), ("Correct Component", acc)]
        if verdict == "Confirmed":
            verdict_reason = "All 3 checks confirmed by evidence."
        elif verdict == "Disputed":
            disputed = [name for name, val in checks if val == "Disputed"]
            verdict_reason = f"Disputed on: {', '.join(disputed)}."
        elif verdict == "Unconfirmed":
            verdict_reason = "Evidence insufficient to confirm any of the 3 checks."
        else:
            verdict_reason = "; ".join(f"{name}: {val}" for name, val in checks) + "."

        _log(f"[audit] Result for '{company_name}' — exists={ace}, supply={ast}, component={acc} → {verdict}")
        return {
            "audit_company_exists":    ace,
            "audit_supply_ties":       ast,
            "audit_correct_component": acc,
            "audit_notes_exists":      str(parsed.get("audit_notes_exists",   "")),
            "audit_notes_supply":      str(parsed.get("audit_notes_supply",   "")),
            "audit_notes_component":   str(parsed.get("audit_notes_component","")),
            "audit_verdict":           verdict,
            "audit_verdict_reason":    verdict_reason,
        }

    _log("[audit] All attempts failed — marking Unconfirmed", True)
    return {
        "audit_company_exists":    "Unconfirmed",
        "audit_supply_ties":       "Unconfirmed",
        "audit_correct_component": "Unconfirmed",
        "audit_notes_exists":      "Audit LLM call failed",
        "audit_notes_supply":      "Audit LLM call failed",
        "audit_notes_component":   "Audit LLM call failed",
        "audit_verdict":           "Unconfirmed",
        "audit_verdict_reason":    "Audit LLM call failed.",
    }


# ═══════════════════════════════════════════════════════════════════════════════
# Standalone Audit Blueprint
# ═══════════════════════════════════════════════════════════════════════════════

_audit_queues:      dict = {}
_audit_queues_lock  = threading.Lock()
_audit_store:       dict = {}   # stream_id → {"path": str, "filename": str}
_audit_store_lock   = threading.Lock()
_audit_cancel:      dict = {}   # stream_id → threading.Event
_audit_cancel_lock  = threading.Lock()


def _audit_push(stream_id: str, event: dict):
    with _audit_queues_lock:
        if stream_id in _audit_queues:
            _audit_queues[stream_id].append(event)


def _find_col(ws, header: str) -> int | None:
    """Return 1-based column index for header (case-insensitive), or None."""
    for cell in ws[1]:
        if cell.value and str(cell.value).strip().lower() == header.lower():
            return cell.column
    return None


def _parse_urls(raw_urls: str) -> list[str]:
    """Split URL column value and strip date suffixes like ' (2024-01-15)'."""
    urls = []
    for entry in str(raw_urls or "").split(", "):
        entry = entry.strip()
        if not entry:
            continue
        clean = entry.rsplit(" (", 1)[0].strip() if " (" in entry else entry
        if clean.startswith("http"):
            urls.append(clean)
    return urls


def _parse_note(verification_notes: str, tag: str) -> str:
    """Extract per-check note from '[Exists] ... | [Supply] ... | [Component] ...' string."""
    for part in str(verification_notes or "").split(" | "):
        part = part.strip()
        if part.startswith(f"[{tag}]"):
            return part[len(f"[{tag}]"):].strip()
    return ""



def _run_audit_standalone(stream_id: str, file_bytes: bytes, filename: str, provider: str):
    """Background thread: read verified Excel, run auditor per row, save result."""

    def push(event):
        _audit_push(stream_id, event)

    def log_fn(msg, is_error=False):
        push({"type": "log", "message": msg, "is_error": is_error})

    _log = _make_logger(log_fn)

    try:
        with _audit_cancel_lock:
            cancel_evt = threading.Event()
            _audit_cancel[stream_id] = cancel_evt

        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        tier_sheets = [ws for ws in wb.worksheets if "tier" in ws.title.lower()]

        if not tier_sheets:
            push({"type": "error", "message": "No tier sheets found in this workbook."})
            return

        # Count rows for progress
        total_rows = 0
        for ws in tier_sheets:
            col = _find_col(ws, "Company Name")
            if col:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[col - 1]:
                        total_rows += 1

        push({"type": "start", "total": total_rows})
        push({"type": "timer_start"})
        processed = 0

        for ws in tier_sheets:
            col_company   = _find_col(ws, "Company Name")
            col_supplies  = _find_col(ws, "Supplies To")
            col_component = _find_col(ws, "Components Supplied")
            col_exists    = _find_col(ws, "Company Exists")
            col_supply    = _find_col(ws, "Supply Ties Exist")
            col_correct   = _find_col(ws, "Correct Component Supplied")
            col_notes     = _find_col(ws, "Verification Notes")
            col_urls      = _find_col(ws, "URLs") or _find_col(ws, "URL")

            if not col_company:
                push({"type": "status", "message": f"⚠️ Skipping '{ws.title}' — no Company Name column"})
                continue

            # Write audit header columns once per sheet
            audit_start_col = ws.max_column + 1
            write_audit_headers(ws, audit_start_col, provider)

            from concurrent.futures import ThreadPoolExecutor, as_completed

            def audit_row(row_idx, company, supplies_to, components, exists_val,
                          supply_val, correct_val, notes_val, urls_val):
                # Re-crawl URLs from the verified Excel to rebuild evidence
                urls    = _parse_urls(urls_val)
                crawled = crawl_urls(urls[:4], log_fn=log_fn) if urls else []
                evidence = "\n\n".join(
                    f"[{r['url']}]\n{r['content']}" for r in crawled
                ) if crawled else ""

                stage1 = {
                    "company_exists":    exists_val == "Yes",
                    "supply_ties":       supply_val == "Yes",
                    "correct_component": correct_val == "Yes",
                    "notes_exists":      _parse_note(notes_val, "Exists"),
                    "notes_supply":      _parse_note(notes_val, "Supply"),
                    "notes_component":   _parse_note(notes_val, "Component"),
                }

                return row_idx, ai_audit_row(
                    evidence, company, supplies_to, components, stage1, provider, log_fn=log_fn
                )

            rows_data = []
            for row_idx in range(2, ws.max_row + 1):
                company = ws.cell(row=row_idx, column=col_company).value
                if not company:
                    continue
                def _get(col):
                    return str(ws.cell(row=row_idx, column=col).value or "") if col else ""
                rows_data.append((
                    row_idx,
                    str(company),
                    _get(col_supplies),
                    _get(col_component),
                    _get(col_exists),
                    _get(col_supply),
                    _get(col_correct),
                    _get(col_notes),
                    _get(col_urls),
                ))

            with ThreadPoolExecutor(max_workers=3) as executor:
                futures = {
                    executor.submit(audit_row, *row): row[0]
                    for row in rows_data
                }
                for future in as_completed(futures):
                    processed += 1
                    try:
                        row_idx, result = future.result()
                        write_audit_row(ws, row_idx, audit_start_col, result)
                        verdict = result.get("audit_verdict", "?")
                        company = ws.cell(row=row_idx, column=col_company).value
                        push({
                            "type": "progress",
                            "processed": processed,
                            "total": total_rows,
                            "message": f"[{provider}] {company} → {verdict}",
                        })
                    except Exception as e:
                        push({"type": "log", "message": f"[error] Row failed: {e}", "is_error": True})

                    with _audit_cancel_lock:
                        evt = _audit_cancel.get(stream_id)
                    if evt and evt.is_set():
                        for f in futures:
                            f.cancel()
                        push({"type": "cancelled", "processed": processed})
                        break

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(buf.read())
            tmp_path = tmp.name

        import datetime
        date_str = datetime.date.today().strftime("%Y%m%d")
        out_filename = f"audited_{date_str}_{filename}"
        with _audit_store_lock:
            _audit_store[stream_id] = {"path": tmp_path, "filename": out_filename}

        push({"type": "timer_stop"})
        push({"type": "complete", "message": "✅ Audit complete — downloading…"})

    except Exception as e:
        push({"type": "error", "message": str(e)})
    finally:
        push({"type": "stream_end"})
        with _audit_cancel_lock:
            _audit_cancel.pop(stream_id, None)


# ── Routes ────────────────────────────────────────────────────────────────────

@audit_bp.route("/dev/audit")
def audit_page():
    html = """<!DOCTYPE html>
<html>
<head>
  <title>Supplier Audit</title>
  <link href="https://fonts.googleapis.com/css2?family=Syne:wght@800&display=swap" rel="stylesheet">
  <style>
    .logo{font-family:'Syne',sans-serif;font-weight:800;font-size:1.4rem;letter-spacing:-.02em}
    .logo span{color:#e8ff47}
    body { font-family: monospace; padding: 2rem; background: #111; color: #eee; }
    h2 span { color: #e8ff47; }
    button { background: #333; color: #eee; border: 1px solid #555; padding: .5rem 1.5rem; cursor: pointer; font-family: monospace; }
    button:hover { background: #444; }
    button:disabled { opacity: .4; cursor: not-allowed; }
    #stopBtn { background: #5c1a1a; border-color: #a33; display: none; margin-left: .5rem; }
    #stopBtn:hover { background: #7a2020; }
    #stopBtn:disabled { opacity: .5; }
    #log-container { height: 400px; overflow-y: auto; background: #1a1a1a; border: 1px solid #333; padding: .75rem; margin-top: 1.5rem; border-radius: 4px; }
    #log div { margin: 2px 0; font-size: .85rem; line-height: 1.6; }
    .done    { color: #4caf50; }
    .error   { color: #f44336; }
    .log-line { color: #888; font-size: .78rem; }
    .log-err  { color: #e57373; font-size: .78rem; }
    select { background: #222; color: #eee; border: 1px solid #555; padding: .3rem; font-family: monospace; margin-bottom: 1rem; }
    label  { font-size: .9rem; color: #aaa; }
    #timer { display: none; font-size: .85rem; color: #e8ff47; margin-top: .75rem; letter-spacing: .05em; }
    #timer span { font-weight: bold; }
    #file-list { margin: .75rem 0; display: flex; flex-direction: column; gap: .35rem; min-height: 1.5rem; }
    .file-row { display: flex; align-items: center; gap: .75rem; padding: .4rem .65rem; background: #1a1a1a; border: 1px solid #333; border-radius: 3px; }
    .file-name { flex: 1; font-size: .83rem; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
    .file-size { color: #555; font-size: .75rem; flex-shrink: 0; }
    .file-status { font-size: .72rem; min-width: 95px; text-align: right; flex-shrink: 0; color: #555; }
    .file-status.done { color: #4caf50; }
    .file-status.error { color: #f44336; }
    .file-status.processing { color: #e8ff47; }
    .remove-btn { background: none; border: none; color: #555; cursor: pointer; padding: 0 2px; font-size: .85rem; line-height: 1; flex-shrink: 0; font-family: monospace; }
    .remove-btn:hover:not(:disabled) { color: #f44336; }
    #addBtn { background: #1a2a1a; color: #4caf50; border: 1px dashed #2e5c2e; padding: .35rem 1rem; font-size: .82rem; font-family: monospace; cursor: pointer; }
    #addBtn:hover:not(:disabled) { background: #213321; }
  </style>
</head>
<body>
  <div class="logo" style="margin-bottom:1.5rem">
    <img src="/static/images/nusW-tliap_transparent_bg.png" width="320" height="80" style="vertical-align:middle"> Supplier<span>Map</span>
  </div>
  <h2>Supplier <span>Audit</span></h2>
  <p style="color:#aaa;font-size:.88rem;max-width:700px;line-height:1.7;margin-bottom:1.2rem">
    Upload verified Excel files from the verification pipeline. The auditor re-crawls the
    source URLs already in the file to rebuild the original evidence, then adversarially
    challenges Stage 1's verdicts — looking for weak inferences, gaps, or overreach.
    Upload files <strong>without</strong> existing audit columns.
  </p>

  <label>Auditor model:</label><br>
  <select id="provider">
    <option value="gemini">Gemini</option>
    <option value="anthropic">Anthropic</option>
  </select>

  <div id="file-list"><div style="color:#555;font-size:.82rem;padding:.3rem 0">No files added.</div></div>
  <input type="file" id="hiddenInput" accept=".xlsx" multiple style="display:none" onchange="addFiles(this.files)">
  <button id="addBtn" onclick="document.getElementById('hiddenInput').click()">+ Add Files</button>

  <div style="margin-top:1rem">
    <button id="uploadBtn" onclick="startAudit()">▶ Upload &amp; Audit</button>
    <button id="stopBtn" onclick="stopAudit()">⛔ Stop</button>
  </div>

  <div id="timer">⏱ Elapsed: <span id="timerVal">0.0s</span></div>
  <div id="log-container"><div id="log"></div></div>

<script>
  let fileQueue = [], activeStreamId = null, isCancelled = false;
  let timerInterval = null, timerStart = null;
  const uploadBtn = document.getElementById('uploadBtn');
  const stopBtn   = document.getElementById('stopBtn');
  const addBtn    = document.getElementById('addBtn');

  function setRunning(v) {
    uploadBtn.disabled = v;
    addBtn.disabled    = v;
    stopBtn.style.display = v ? 'inline-block' : 'none';
    stopBtn.disabled      = false;
    stopBtn.textContent   = '⛔ Stop';
    document.querySelectorAll('.remove-btn').forEach(b => b.disabled = v);
  }

  function startTimer() {
    timerStart = Date.now();
    document.getElementById('timer').style.display = 'block';
    timerInterval = setInterval(() => {
      document.getElementById('timerVal').textContent = ((Date.now() - timerStart) / 1000).toFixed(1) + 's';
    }, 100);
  }
  function stopTimer() {
    clearInterval(timerInterval);
  }

  function append(msg, cls) {
    const d = document.getElementById('log');
    const el = document.createElement('div');
    if (cls) el.className = cls;
    el.textContent = msg;
    d.appendChild(el);
    d.parentElement.scrollTop = d.parentElement.scrollHeight;
  }

  function renderFileList() {
    const el = document.getElementById('file-list');
    if (!fileQueue.length) {
      el.innerHTML = '<div style="color:#555;font-size:.82rem;padding:.3rem 0">No files added.</div>';
      return;
    }
    el.innerHTML = fileQueue.map((f, i) => `
      <div class="file-row" id="file-row-${i}">
        <span class="file-name">${f.name}</span>
        <span class="file-size">${(f.size/1024).toFixed(1)} KB</span>
        <span class="file-status" id="file-status-${i}">queued</span>
        <button class="remove-btn" onclick="removeFile(${i})">✕</button>
      </div>`).join('');
  }

  function addFiles(files) {
    for (const f of files) fileQueue.push(f);
    document.getElementById('hiddenInput').value = '';
    renderFileList();
  }
  function removeFile(i) {
    fileQueue.splice(i, 1);
    renderFileList();
  }
  function setFileStatus(i, text, cls) {
    const el = document.getElementById('file-status-' + i);
    if (!el) return;
    el.textContent = text;
    el.className = 'file-status' + (cls ? ' ' + cls : '');
  }

  async function processFile(file, idx, provider) {
    setFileStatus(idx, 'uploading…', 'processing');
    const fd = new FormData();
    fd.append('file', file);
    fd.append('provider', provider);

    let res;
    try {
      res = await fetch('/api/audit/upload', { method: 'POST', body: fd });
    } catch(e) {
      append('[' + file.name + '] Upload failed: ' + e, 'error');
      setFileStatus(idx, '✗ failed', 'error');
      return false;
    }
    if (!res.ok) {
      append('[' + file.name + '] Server error', 'error');
      setFileStatus(idx, '✗ failed', 'error');
      return false;
    }
    const { stream_id } = await res.json();
    activeStreamId = stream_id;
    setFileStatus(idx, 'auditing…', 'processing');

    return new Promise(resolve => {
      const es = new EventSource('/api/audit/stream/' + stream_id);
      es.onmessage = e => {
        const msg = JSON.parse(e.data);
        if (msg.type === 'log') {
          append(msg.message, msg.is_error ? 'log-err' : 'log-line');
        } else if (msg.type === 'progress') {
          append('✓ ' + msg.message, '');
        } else if (msg.type === 'start') {
          append('[' + file.name + '] Starting audit — ' + msg.total + ' row(s)', '');
        } else if (msg.type === 'timer_start') {
          startTimer();
        } else if (msg.type === 'timer_stop') {
          stopTimer();
        } else if (msg.type === 'complete') {
          append(msg.message, 'done');
          setFileStatus(idx, '✓ done', 'done');
          window.location.href = '/api/audit/download/' + stream_id;
          es.close(); resolve(true);
        } else if (msg.type === 'cancelled') {
          append('⛔ Audit stopped — downloading partial results…', 'error');
          setFileStatus(idx, '⛔ stopped', 'error');
          window.location.href = '/api/audit/download/' + stream_id;
          es.close(); resolve(false);
        } else if (msg.type === 'error') {
          append('Error: ' + msg.message, 'error');
          setFileStatus(idx, '✗ error', 'error');
          es.close(); resolve(false);
        } else if (msg.type === 'stream_end') {
          es.close(); resolve(false);
        }
      };
      es.onerror = () => { es.close(); resolve(false); };
    });
  }

  async function startAudit() {
    if (!fileQueue.length) { alert('Add at least one .xlsx file.'); return; }
    document.getElementById('log').innerHTML = '';
    isCancelled = false;
    setRunning(true);
    const provider = document.getElementById('provider').value;
    let ok = 0;
    for (let i = 0; i < fileQueue.length; i++) {
      if (isCancelled) { append('⛔ Queue stopped.', 'error'); break; }
      if (await processFile(fileQueue[i], i, provider)) ok++;
    }
    stopTimer();
    if (!isCancelled) append('Done — ' + ok + '/' + fileQueue.length + ' file(s) audited.', 'done');
    setRunning(false);
  }

  async function stopAudit() {
    isCancelled = true;
    stopBtn.disabled = true;
    stopBtn.textContent = '⛔ Stopping…';
    if (activeStreamId) {
      try { await fetch('/api/audit/cancel/' + activeStreamId, { method: 'POST' }); } catch(e) {}
    }
  }
</script>
</body>
</html>"""
    return html


@audit_bp.route("/api/audit/upload", methods=["POST"])
def audit_upload():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    if not f.filename.endswith(".xlsx"):
        return jsonify({"error": "Only .xlsx files are supported"}), 400

    provider   = request.form.get("provider", "gemini")
    file_bytes = f.read()
    filename   = f.filename
    stream_id  = uuid.uuid4().hex

    with _audit_queues_lock:
        _audit_queues[stream_id] = []

    thread = threading.Thread(
        target=_run_audit_standalone,
        args=(stream_id, file_bytes, filename, provider),
        daemon=True,
    )
    thread.start()
    return jsonify({"stream_id": stream_id})


@audit_bp.route("/api/audit/stream/<stream_id>")
def audit_stream(stream_id):
    def generate():
        while True:
            with _audit_queues_lock:
                queue = _audit_queues.get(stream_id, [])
                events, _audit_queues[stream_id] = queue[:], []
            for event in events:
                import json
                yield f"data: {json.dumps(event)}\n\n"
                if event.get("type") == "stream_end":
                    return
            time.sleep(0.3)
    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@audit_bp.route("/api/audit/download/<stream_id>")
def audit_download(stream_id):
    with _audit_store_lock:
        info = _audit_store.get(stream_id)
    if not info:
        return jsonify({"error": "File not ready or stream_id invalid"}), 404
    from flask import send_file
    return send_file(info["path"], as_attachment=True, download_name=info["filename"],
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@audit_bp.route("/api/audit/cancel/<stream_id>", methods=["POST"])
def audit_cancel(stream_id):
    with _audit_cancel_lock:
        evt = _audit_cancel.get(stream_id)
    if evt:
        evt.set()
        return jsonify({"ok": True})
    return jsonify({"error": "No active job"}), 404
