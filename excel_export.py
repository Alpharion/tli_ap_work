"""
excel_export.py — Excel workbook generation for supply chain exports.

Exports:
  build_bulk_workbook(results, provider, depth) -> bytes
"""

import datetime
import io

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Styles ─────────────────────────────────────────────────────────────────────

FONT = "Arial"

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=9, color="000000"):
    return Font(name=FONT, bold=bold, size=size, color=color)

def _align(wrap=True):
    return Alignment(horizontal="left", vertical="top", wrap_text=wrap)

def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

HDR_PRODUCT = _fill("1A237E")
HDR_OEM     = _fill("1B5E20")
HDR_TIER    = [
    _fill("1565C0"),
    _fill("6A1B9A"),
    _fill("E65100"),
    _fill("37474F"),
]
ALT_ROW   = _fill("F5F5F5")
WHITE_ROW = _fill("FFFFFF")


# ── Cell helpers ───────────────────────────────────────────────────────────────

def set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def hdr_cell(ws, row, col, value, fill):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill
    c.font = _font(bold=True, size=9, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = _border()
    return c

def data_cell(ws, row, col, value, fill=None, bold=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(bold=bold)
    c.alignment = _align()
    c.border = _border()
    if fill:
        c.fill = fill
    return c

def _row_fill(idx):
    return ALT_ROW if idx % 2 == 0 else WHITE_ROW


# ── Sheet writers ──────────────────────────────────────────────────────────────

def write_index_sheet(ws, results: list[dict], run_meta: dict):
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 45

    meta_rows = [
        ("Run Date",                  run_meta["run_date"]),
        ("Provider",                  run_meta["provider"]),
        ("Depth",                     str(run_meta["depth"])),
        ("Products Run",              str(run_meta["total_products"])),
        ("Total Suppliers Collected", str(run_meta["total_suppliers"])),
    ]
    title = ws.cell(row=1, column=1, value="Bulk Export - Run Summary")
    title.font = Font(name=FONT, bold=True, size=14, color="1A237E")
    ws.merge_cells("A1:B1")

    for r_idx, (k, v) in enumerate(meta_rows, 3):
        data_cell(ws, r_idx, 1, k, bold=True, fill=_fill("E8EAF6"))
        data_cell(ws, r_idx, 2, v)

    tbl_start = len(meta_rows) + 5
    headers = ["Product Input", "Identified Name", "Industry",
               "OEMs Found", "T1 Suppliers", "T2 Suppliers", "T3 Suppliers", "Sheet Name"]
    col_widths = [28, 28, 18, 10, 12, 12, 12, 22]
    set_col_widths(ws, col_widths)

    for ci, h in enumerate(headers, 1):
        hdr_cell(ws, tbl_start, ci, h, HDR_PRODUCT)
    ws.row_dimensions[tbl_start].height = 24

    for ri, r in enumerate(results, tbl_start + 1):
        sc   = r.get("supply_chain", {})
        fill = _row_fill(ri)

        product = sc.get("product", {})
        if not product.get("product_name") or product.get("industry", "Unknown") == "Unknown":
            for prov_sc in r.get("provider_results", {}).values():
                p = prov_sc.get("product", {})
                if p.get("product_name") and p.get("industry") != "Unknown":
                    product = p
                    break

        data_cell(ws, ri, 1, r["product_input"],                                         fill=fill, bold=True)
        data_cell(ws, ri, 2, sc.get("product", {}).get("product_name", "—"),             fill=fill)
        data_cell(ws, ri, 3, sc.get("product", {}).get("industry", "—"),                 fill=fill)
        data_cell(ws, ri, 4, len(sc.get("oems", [])),                                    fill=fill)
        data_cell(ws, ri, 5, len(sc.get("tiers", {}).get("tier_1", [])),                 fill=fill)
        data_cell(ws, ri, 6, len(sc.get("tiers", {}).get("tier_2", [])),                 fill=fill)
        data_cell(ws, ri, 7, len(sc.get("tiers", {}).get("tier_3", [])),                 fill=fill)
        data_cell(ws, ri, 8, r.get("sheet_prefix", ""),                                  fill=fill)

    ws.freeze_panes = f"A{tbl_start + 1}"
    ws.auto_filter.ref = (
        f"A{tbl_start}:{get_column_letter(len(headers))}{tbl_start + len(results)}"
    )


HDR_REG = _fill("4A148C")

REG_STATUS_FILL = {
    "approved":  _fill("C8E6C9"),
    "registered": _fill("C8E6C9"),
    "pending":   _fill("FFF9C4"),
    "not found": _fill("FFCDD2"),
    "unknown":   _fill("F5F5F5"),
}


def write_oem_sheet(ws, results: list[dict]):
    base_headers = ["Product", "Company Name", "Country", "Role", "Market Share",
                    "Market Share %", "Market Share Source", "Market Rank",
                    "Confidence", "Notes", "AI Provider", "Flagged"]
    reg_headers  = ["Regulatory Body", "Region", "Status", "Details", "Confidence (Reg)"]
    headers      = base_headers + reg_headers
    col_widths   = [28, 26, 14, 22, 12, 14, 28, 12, 11, 35, 14, 10, 18, 14, 14, 40, 12]
    set_col_widths(ws, col_widths)

    for ci, h in enumerate(base_headers, 1):
        hdr_cell(ws, 1, ci, h, HDR_OEM)
    for ci, h in enumerate(reg_headers, len(base_headers) + 1):
        hdr_cell(ws, 1, ci, h, HDR_REG)
    ws.row_dimensions[1].height = 24

    row = 2
    for r in results:
        for prov_id, prov_sc in r.get("provider_results", {r.get("provider", "unknown"): r.get("supply_chain", {})}).items():
            product_name = prov_sc.get("product", {}).get("product_name", r["product_input"])

            # build lookup: company_name (lowercase) → list of regulatory body dicts
            reg_lookup = {}
            for entry in prov_sc.get("regulatory", []):
                key = entry.get("company_name", "").strip().lower()
                if key:
                    reg_lookup[key] = entry.get("regulatory_bodies", [])

            for oem in prov_sc.get("oems", []):
                fill = _row_fill(row)
                rank = oem.get("market_rank", "")

                oem_name   = oem.get("company_name", "")
                reg_bodies = reg_lookup.get(oem_name.strip().lower(), [])

                def _write_oem_base(r_idx, fill):
                    data_cell(ws, r_idx, 1,  product_name,                        bold=True, fill=fill)
                    data_cell(ws, r_idx, 2,  oem_name,                            fill=fill)
                    data_cell(ws, r_idx, 3,  oem.get("country", ""),              fill=fill)
                    data_cell(ws, r_idx, 4,  oem.get("role", ""),                 fill=fill)
                    data_cell(ws, r_idx, 5,  oem.get("market_share", ""),         fill=fill)
                    data_cell(ws, r_idx, 6,  oem.get("market_share_pct", ""),     fill=fill)
                    data_cell(ws, r_idx, 7,  oem.get("market_share_source", ""),  fill=fill)
                    data_cell(ws, r_idx, 8,  "" if rank == 99 else rank,          fill=fill)
                    data_cell(ws, r_idx, 9,  oem.get("confidence", ""),           fill=fill)
                    data_cell(ws, r_idx, 10, oem.get("notes", ""),                fill=fill)
                    data_cell(ws, r_idx, 11, prov_id,                             fill=fill)
                    data_cell(ws, r_idx, 12, str(oem.get("flagged", False)),      fill=fill)

                if not reg_bodies:
                    # non-pharma or no regulatory data — single row, reg columns blank
                    _write_oem_base(row, fill)
                    row += 1
                else:
                    # pharma — one row per regulatory body
                    for body in reg_bodies:
                        status   = body.get("status", "unknown").lower()
                        reg_fill = REG_STATUS_FILL.get(status, REG_STATUS_FILL["unknown"])
                        _write_oem_base(row, fill)
                        data_cell(ws, row, 13, body.get("body", ""),       fill=reg_fill)
                        data_cell(ws, row, 14, body.get("country", ""),    fill=reg_fill)
                        data_cell(ws, row, 15, body.get("status", ""),     fill=reg_fill)
                        data_cell(ws, row, 16, body.get("details", ""),    fill=reg_fill)
                        data_cell(ws, row, 17, body.get("confidence", ""), fill=reg_fill)
                        row += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row - 1}"


def write_tier_sheet(ws, results: list[dict], tier_key: str, tier_num: int):
    base_headers   = ["Product", "Company Name", "Country", "Supplies To", "OEM Root",
                      "Components Supplied", "Confidence", "Source Hint", "AI Provider", "Flagged"]
    verify_headers = ["Verification Notes", "Company Exists", "Supply Ties Exist",
                      "Correct Component Supplied", "URL"]
    headers    = base_headers + verify_headers
    col_widths = [28, 26, 14, 22, 20, 32, 11, 35, 14, 10, 40, 14, 14, 14, 50]
    set_col_widths(ws, col_widths)

    hdr_fill    = HDR_TIER[tier_num - 1] if tier_num - 1 < len(HDR_TIER) else _fill("37474F")
    verify_fill = _fill("E65100")
    for ci, h in enumerate(base_headers, 1):
        hdr_cell(ws, 1, ci, h, hdr_fill)
    for ci, h in enumerate(verify_headers, len(base_headers) + 1):
        hdr_cell(ws, 1, ci, h, verify_fill)
    ws.row_dimensions[1].height = 24

    row = 2
    for r in results:
        for prov_id, prov_sc in r.get("provider_results", {r.get("provider", "unknown"): r.get("supply_chain", {})}).items():
            product_name = prov_sc.get("product", {}).get("product_name", r["product_input"])
            for s in prov_sc.get("tiers", {}).get(tier_key, []):
                fill       = _row_fill(row)
                components = ", ".join(s.get("components_supplied") or [])
                data_cell(ws, row, 1,  product_name,                    bold=True, fill=fill)
                data_cell(ws, row, 2,  s.get("company_name", ""),       fill=fill)
                data_cell(ws, row, 3,  s.get("country", ""),            fill=fill)
                data_cell(ws, row, 4,  s.get("supplies_to", ""),        fill=fill)
                data_cell(ws, row, 5,  s.get("oem_root", ""),           fill=fill)
                data_cell(ws, row, 6,  components,                      fill=fill)
                data_cell(ws, row, 7,  s.get("confidence", ""),         fill=fill)
                data_cell(ws, row, 8,  s.get("source_hint", ""),        fill=fill)
                data_cell(ws, row, 9,  prov_id,                         fill=fill)
                data_cell(ws, row, 10, str(s.get("flagged", False)),    fill=fill)
                # cols 11-15: verification columns — filled by verify layer
                row += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row - 1}"


def write_product_sheet(ws, result: dict):
    provider_results = result.get("provider_results", {})
    if not provider_results:
        sc = result.get("supply_chain", {})
        provider_results = {sc.get("provider", "unknown"): sc}

    product = {}
    for sc in provider_results.values():
        p = sc.get("product", {})
        if p.get("product_name") and p.get("industry") != "Unknown":
            product = p
            break

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 70

    title = ws.cell(row=1, column=1, value=product.get("product_name", result["product_input"]))
    title.font = Font(name=FONT, bold=True, size=13, color="1A237E")
    ws.merge_cells("A1:B1")

    overview_fields = [
        ("Category",         product.get("product_category", "")),
        ("Industry",         product.get("industry", "")),
        ("OEM Manufacturer", product.get("oem_manufacturer", "")),
        ("OEM Country",      product.get("oem_country", "")),
        ("Key Components",   ", ".join(product.get("key_components") or [])),
        ("Description",      product.get("description", "")),
    ]
    for r_idx, (label, value) in enumerate(overview_fields, 3):
        data_cell(ws, r_idx, 1, label, bold=True, fill=_fill("E8EAF6"))
        data_cell(ws, r_idx, 2, value)
        ws.row_dimensions[r_idx].height = 90

    current_row = len(overview_fields) + 5
    tier_headers = ["Company Name", "Country", "Supplies To", "OEM Root",
                    "Components", "Confidence", "Source Hint", "Source URL"]
    tier_widths  = [28, 14, 22, 20, 32, 11, 35, 50]

    for prov_id, sc in provider_results.items():
        prov_heading = ws.cell(row=current_row, column=1, value=f"Provider: {prov_id.upper()}")
        prov_heading.font = Font(name=FONT, bold=True, size=12, color="FFFFFF")
        prov_heading.fill = _fill("37474F")
        ws.merge_cells(f"A{current_row}:H{current_row}")
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        summary_cell = ws.cell(row=current_row, column=1, value=sc.get("summary", "No summary generated."))
        summary_cell.font = _font()
        summary_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.merge_cells(f"A{current_row}:H{current_row}")
        ws.row_dimensions[current_row].height = 50
        current_row += 2

        for ti, (tier_key, suppliers) in enumerate(sc.get("tiers", {}).items()):
            hdr_fill   = HDR_TIER[ti] if ti < len(HDR_TIER) else _fill("37474F")
            tier_label = tier_key.replace("_", " ").upper()

            if ti == 0 and prov_id == next(iter(provider_results)):
                for ci, w in enumerate(tier_widths, 1):
                    ws.column_dimensions[get_column_letter(ci)].width = w

            heading = ws.cell(row=current_row, column=1,
                               value=f"{tier_label} ({len(suppliers)} suppliers)")
            heading.font = Font(name=FONT, bold=True, size=11, color="FFFFFF")
            heading.fill = hdr_fill
            ws.merge_cells(f"A{current_row}:H{current_row}")
            ws.row_dimensions[current_row].height = 22
            current_row += 1

            for ci, h in enumerate(tier_headers, 1):
                hdr_cell(ws, current_row, ci, h, hdr_fill)
            ws.row_dimensions[current_row].height = 20
            current_row += 1

            if not suppliers:
                ws.cell(row=current_row, column=1, value="No suppliers identified")
                current_row += 2
                continue

            for s in suppliers:
                fill = _row_fill(current_row)
                data_cell(ws, current_row, 1, s.get("company_name", ""), bold=True, fill=fill)
                data_cell(ws, current_row, 2, s.get("country", ""),       fill=fill)
                data_cell(ws, current_row, 3, s.get("supplies_to", ""),   fill=fill)
                data_cell(ws, current_row, 4, s.get("oem_root", ""),      fill=fill)
                data_cell(ws, current_row, 5, ", ".join(s.get("components_supplied") or []), fill=fill)
                data_cell(ws, current_row, 6, s.get("confidence", ""),    fill=fill)
                data_cell(ws, current_row, 7, s.get("source_hint", ""),   fill=fill)
                data_cell(ws, current_row, 8, s.get("source_url", ""),    fill=fill)
                ws.row_dimensions[current_row].height = 60
                current_row += 1

            current_row += 1

        current_row += 2


# ── Workbook builder ───────────────────────────────────────────────────────────

def build_bulk_workbook(results: list[dict], provider: str, depth: int) -> bytes:
    wb = openpyxl.Workbook()
    total_suppliers = sum(
        sum(len(r["supply_chain"].get("tiers", {}).get(f"tier_{t}", [])) for t in range(1, 4))
        for r in results
    )

    run_meta = {
        "run_date":        datetime.datetime.now().strftime("%Y-%m-%d %H:%M UTC"),
        "provider":        provider,
        "depth":           depth,
        "total_products":  len(results),
        "total_suppliers": total_suppliers,
    }

    ws_index = wb.active
    ws_index.title = "INDEX"
    write_index_sheet(ws_index, results, run_meta)

    ws_oem = wb.create_sheet("ALL_OEMs")
    write_oem_sheet(ws_oem, results)

    for tier_num in range(1, depth + 1):
        tier_key = f"tier_{tier_num}"
        ws_tier  = wb.create_sheet(f"ALL_TIER_{tier_num}")
        write_tier_sheet(ws_tier, results, tier_key, tier_num)

    for r in results:
        ws_prod = wb.create_sheet(r["sheet_prefix"])
        write_product_sheet(ws_prod, r)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
