"""
reg_crawl.py — Regulatory-body webcrawl layer.

For each row in the ALL_OEMs sheet that has both a Company Name and a
Regulatory Body, this layer:
  1. Searches DuckDuckGo for "<company> <regulatory_body> regulatory approval"
  2. Crawls the top 4 URLs with crawl4ai
  3. Sends the evidence to Gemini and extracts 4 structured fields
  4. Writes those fields as new columns back into ALL_OEMs

New columns appended:
  Webcrawl Reg | Webcrawl Reg Region | Webcrawl Reg Status | Webcrawl Reg Details

Register in app.py via:
    from reg_crawl import reg_crawl_bp
    app.register_blueprint(reg_crawl_bp)
"""

import datetime
import io
import json
import tempfile
import threading
import time
import uuid

import openpyxl
from flask import Blueprint, Response, jsonify, request, send_file
from openpyxl.styles import Alignment, Font, PatternFill

from ai import call_ai, safe_parse_json
from scraper import crawl_urls, ddg_search

reg_crawl_bp = Blueprint("reg_crawl", __name__)

# ── Styling ───────────────────────────────────────────────────────────────────

_HDR_FILL  = PatternFill("solid", fgColor="1A237E")   # dark blue header
_HDR_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
_WRAP      = Alignment(wrap_text=True, vertical="top")

_STATUS_FILLS = {
    "approved":    PatternFill("solid", fgColor="C8E6C9"),  # green
    "pending":     PatternFill("solid", fgColor="FFF9C4"),  # yellow
    "not found":   PatternFill("solid", fgColor="FFCDD2"),  # red
    "expired":     PatternFill("solid", fgColor="FFE0B2"),  # orange
    "unknown":     PatternFill("solid", fgColor="F5F5F5"),  # grey
}

_NEW_HEADERS = [
    ("Webcrawl Reg",        28),
    ("Webcrawl Reg Region", 18),
    ("Webcrawl Reg Status", 16),
    ("Webcrawl Reg Details", 55),
    ("Webcrawl Reg URLs",   60),
]

# ── Shared state ──────────────────────────────────────────────────────────────

_queues:      dict = {}
_queues_lock  = threading.Lock()
_store:       dict = {}
_store_lock   = threading.Lock()
_cancel:      dict = {}
_cancel_lock  = threading.Lock()


def _push(stream_id: str, event: dict):
    with _queues_lock:
        if stream_id in _queues:
            _queues[stream_id].append(event)


def _find_col(ws, header: str) -> int | None:
    for cell in ws[1]:
        if cell.value and str(cell.value).strip().lower() == header.lower():
            return cell.column
    return None


# ── AI call ───────────────────────────────────────────────────────────────────

_PROMPT_TMPL = """\
You are a pharmaceutical regulatory intelligence analyst.

COMPANY: {company}
REGULATORY BODY: {reg_body}

WEB EVIDENCE (crawled from top search results):
{evidence}

Based solely on the evidence above, fill in the JSON below.
If the evidence does not mention this company or regulatory body, set webcrawl_reg_status to "Not Found".

Return ONLY valid JSON — no markdown, no extra text:
{{
  "webcrawl_reg":        "<the exact regulatory body name as confirmed on the web, or blank if not found>",
  "webcrawl_reg_region": "<jurisdiction / country / region of the approval, e.g. USA, EU, Japan>",
  "webcrawl_reg_status": "<one of: Approved, Pending, Expired, Not Found, Unknown>",
  "webcrawl_reg_details": "<1-2 sentence factual summary of what the evidence says about this company's regulatory status with this body>"
}}"""


def _query_gemini(company: str, reg_body: str, evidence: str, provider: str, log) -> dict:
    prompt = _PROMPT_TMPL.format(
        company=company, reg_body=reg_body,
        evidence=evidence[:4000] if evidence else "(no evidence found)",
    )
    for attempt in range(1, 4):
        try:
            raw    = call_ai(prompt, provider, max_tokens=512)
            parsed = safe_parse_json(raw)
            if isinstance(parsed, dict) and "webcrawl_reg_status" in parsed:
                return parsed
            log(f"[llm] Attempt {attempt}: missing key — retrying")
        except Exception as e:
            log(f"[llm] Attempt {attempt} error: {e}")
        if attempt < 3:
            time.sleep(4 * attempt)
    return {
        "webcrawl_reg":        "",
        "webcrawl_reg_region": "",
        "webcrawl_reg_status": "Unknown",
        "webcrawl_reg_details": "LLM call failed after 3 attempts.",
    }


# ── Background worker ─────────────────────────────────────────────────────────

def _run_reg_crawl(stream_id: str, file_bytes: bytes, filename: str, provider: str):

    def push(event):
        _push(stream_id, event)

    def log(msg, is_error=False):
        push({"type": "log", "message": msg, "is_error": is_error})

    try:
        with _cancel_lock:
            cancel_evt = threading.Event()
            _cancel[stream_id] = cancel_evt

        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))

        # Find the ALL_OEMs sheet
        oem_ws = None
        for ws in wb.worksheets:
            if "oem" in ws.title.lower():
                oem_ws = ws
                break

        if oem_ws is None:
            push({"type": "error", "message": "No OEM sheet found in this workbook."})
            return

        col_company = _find_col(oem_ws, "Company Name")
        col_reg     = _find_col(oem_ws, "Regulatory Body")
        col_product = _find_col(oem_ws, "Product")

        if not col_company or not col_reg:
            push({
                "type": "error",
                "message": (
                    "ALL_OEMs sheet is missing 'Company Name' or 'Regulatory Body' column. "
                    "Make sure you upload a pharma-mode Excel."
                ),
            })
            return

        # Count eligible rows
        eligible = [
            r for r in range(2, oem_ws.max_row + 1)
            if oem_ws.cell(row=r, column=col_company).value
            and oem_ws.cell(row=r, column=col_reg).value
        ]
        push({"type": "start", "total": len(eligible)})
        push({"type": "timer_start"})
        log(f"Sheet '{oem_ws.title}': {len(eligible)} row(s) with Company Name + Regulatory Body")

        # Write new header columns
        start_col = oem_ws.max_column + 1
        for i, (hdr, width) in enumerate(_NEW_HEADERS):
            cell = oem_ws.cell(row=1, column=start_col + i, value=hdr)
            cell.fill = _HDR_FILL
            cell.font = _HDR_FONT
            oem_ws.column_dimensions[cell.column_letter].width = width

        # Collect row data upfront so threads don't read the worksheet concurrently
        rows_data = [
            (
                row_idx,
                str(oem_ws.cell(row=row_idx, column=col_company).value or "").strip(),
                str(oem_ws.cell(row=row_idx, column=col_reg).value or "").strip(),
                str(oem_ws.cell(row=row_idx, column=col_product).value or "").strip() if col_product else "",
            )
            for row_idx in eligible
        ]

        def process_row(row_idx, company, reg_body, product):
            parts = [p for p in [company, product, reg_body, "regulatory approval"] if p]
            query = " ".join(parts)
            log(f"[{row_idx}] Searching: {query}")
            try:
                urls = ddg_search(query, n=6, log_fn=log)
            except Exception as e:
                log(f"[{row_idx}] DDG search failed: {e}", is_error=True)
                urls = []
            crawled  = crawl_urls(urls[:4], log_fn=log) if urls else []
            evidence = "\n\n".join(f"[{r['url']}]\n{r['content']}" for r in crawled) if crawled else ""
            log(f"[{row_idx}] Crawled {len(crawled)} page(s) — calling {provider}…")
            result   = _query_gemini(company, reg_body, evidence, provider, log)
            url_list = ", ".join(
                f"{r['url']} ({r['timestamp']})" if r.get("timestamp") and r["timestamp"] != "unknown"
                else r["url"]
                for r in crawled
            )
            return row_idx, company, reg_body, result, url_list

        from concurrent.futures import ThreadPoolExecutor, as_completed

        processed = 0
        with ThreadPoolExecutor(max_workers=3) as executor:
            futures = {
                executor.submit(process_row, *row): row[0]
                for row in rows_data
            }
            for future in as_completed(futures):
                with _cancel_lock:
                    evt = _cancel.get(stream_id)
                if evt and evt.is_set():
                    for f in futures:
                        f.cancel()
                    push({"type": "cancelled", "processed": processed})
                    break

                try:
                    row_idx, company, reg_body, result, url_list = future.result()
                except Exception as e:
                    push({"type": "log", "message": f"[error] Row failed: {e}", "is_error": True})
                    continue

                status_key  = result.get("webcrawl_reg_status", "Unknown").lower()
                status_fill = _STATUS_FILLS.get(status_key, _STATUS_FILLS["unknown"])

                # Write the 5 new columns (worksheet writes must be single-threaded)
                for i, key in enumerate([
                    "webcrawl_reg", "webcrawl_reg_region",
                    "webcrawl_reg_status", "webcrawl_reg_details",
                ]):
                    cell = oem_ws.cell(row=row_idx, column=start_col + i, value=result.get(key, ""))
                    cell.alignment = _WRAP
                    if key == "webcrawl_reg_status":
                        cell.fill = status_fill
                url_cell = oem_ws.cell(row=row_idx, column=start_col + 4, value=url_list)
                url_cell.alignment = _WRAP

                processed += 1
                push({
                    "type":      "progress",
                    "processed": processed,
                    "total":     len(eligible),
                    "message":   f"{company} / {reg_body} → {result.get('webcrawl_reg_status', '?')}",
                })

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(buf.read())
            tmp_path = tmp.name

        date_str     = datetime.date.today().strftime("%Y%m%d")
        out_filename = f"reg_crawl_{date_str}_{filename}"

        with _store_lock:
            _store[stream_id] = {"path": tmp_path, "filename": out_filename}

        push({"type": "timer_stop"})
        push({"type": "complete", "message": "✅ Regulatory crawl complete — downloading…"})

    except Exception as e:
        push({"type": "error", "message": str(e)})
    finally:
        push({"type": "stream_end"})
        with _cancel_lock:
            _cancel.pop(stream_id, None)


# ── Routes ─────────────────────────────────────────────────────────────────────

@reg_crawl_bp.route("/dev/reg-crawl")
def reg_crawl_page():
    html = """<!DOCTYPE html>
<html>
<head>
  <title>Regulatory Webcrawl</title>
  <link href="https://fonts.googleapis.com/css2?family=Syne:wght@800&display=swap" rel="stylesheet">
  <style>
    .logo{font-family:'Syne',sans-serif;font-weight:800;font-size:1.4rem;letter-spacing:-.02em}
    .logo span{color:#e8ff47}
    body { font-family: monospace; padding: 2rem; background: #111; color: #eee; }
    h2 span { color: #e8ff47; }
    button { background: #333; color: #eee; border: 1px solid #555; padding: .5rem 1.5rem; cursor: pointer; font-family: monospace; }
    button:hover { background: #444; }
    button:disabled { opacity: .4; cursor: not-allowed; }
    #stopBtn { background: #5c1a1a; border-color: #a33; display: none; margin-left: .5rem; }
    #stopBtn:hover { background: #7a2020; }
    #stopBtn:disabled { opacity: .5; }
    #log-container { height: 420px; overflow-y: auto; background: #1a1a1a; border: 1px solid #333; padding: .75rem; margin-top: 1.5rem; border-radius: 4px; }
    #log div { margin: 2px 0; font-size: .85rem; line-height: 1.6; }
    .done    { color: #4caf50; }
    .error   { color: #f44336; }
    .log-line { color: #888; font-size: .78rem; }
    .log-err  { color: #e57373; font-size: .78rem; }
    select { background: #222; color: #eee; border: 1px solid #555; padding: .3rem; font-family: monospace; margin-bottom: 1rem; }
    label  { font-size: .9rem; color: #aaa; }
    #timer { display: none; font-size: .85rem; color: #e8ff47; margin-top: .75rem; letter-spacing: .05em; }
    #timer span { font-weight: bold; }
    #file-list { margin: .75rem 0; display: flex; flex-direction: column; gap: .35rem; min-height: 1.5rem; }
    .file-row { display: flex; align-items: center; gap: .75rem; padding: .4rem .65rem; background: #1a1a1a; border: 1px solid #333; border-radius: 3px; }
    .file-name { flex: 1; font-size: .83rem; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
    .file-size { color: #555; font-size: .75rem; flex-shrink: 0; }
    .file-status { font-size: .72rem; min-width: 95px; text-align: right; flex-shrink: 0; color: #555; }
    .file-status.done { color: #4caf50; }
    .file-status.error { color: #f44336; }
    .file-status.processing { color: #e8ff47; }
    .remove-btn { background: none; border: none; color: #555; cursor: pointer; padding: 0 2px; font-size: .85rem; line-height: 1; flex-shrink: 0; font-family: monospace; }
    .remove-btn:hover:not(:disabled) { color: #f44336; }
    #addBtn { background: #1a2a1a; color: #4caf50; border: 1px dashed #2e5c2e; padding: .35rem 1rem; font-size: .82rem; font-family: monospace; cursor: pointer; }
    #addBtn:hover:not(:disabled) { background: #213321; }
    .badge { display: inline-block; font-size: .68rem; padding: 1px 6px; border-radius: 3px; margin-left: .4rem; vertical-align: middle; }
    .badge-new { background: #1A237E; color: #fff; }
  </style>
</head>
<body>
  <div class="logo" style="margin-bottom:1.5rem">
    <img src="/static/images/nusW-tliap_transparent_bg.png" width="320" height="80" style="vertical-align:middle"> Supplier<span>Map</span>
  </div>
  <h2>Regulatory <span>Webcrawl</span></h2>
  <p style="color:#aaa;font-size:.88rem;max-width:700px;line-height:1.7;margin-bottom:1.2rem">
    Upload Excel files exported from the pipeline in <strong>pharma mode</strong>.
    The ALL_OEMs sheet must contain <code>Company Name</code> and <code>Regulatory Body</code> columns
    (one row per OEM × regulatory body). For each row, this layer searches DuckDuckGo, crawls
    the top 4 pages, and asks the selected model to extract four new columns:
    <span class="badge badge-new">Webcrawl Reg</span>
    <span class="badge badge-new">Webcrawl Reg Region</span>
    <span class="badge badge-new">Webcrawl Reg Status</span>
    <span class="badge badge-new">Webcrawl Reg Details</span>
  </p>

  <label>LLM for extraction:</label><br>
  <select id="provider">
    <option value="gemini">Gemini (recommended)</option>
    <option value="anthropic">Anthropic</option>
  </select>

  <div id="file-list"><div style="color:#555;font-size:.82rem;padding:.3rem 0">No files added.</div></div>
  <input type="file" id="hiddenInput" accept=".xlsx" multiple style="display:none" onchange="addFiles(this.files)">
  <button id="addBtn" onclick="document.getElementById('hiddenInput').click()">+ Add Files</button>

  <div style="margin-top:1rem">
    <button id="uploadBtn" onclick="startCrawl()">▶ Upload &amp; Crawl</button>
    <button id="stopBtn" onclick="stopCrawl()">⛔ Stop</button>
  </div>

  <div id="timer">⏱ Elapsed: <span id="timerVal">0.0s</span></div>
  <div id="log-container"><div id="log"></div></div>

<script>
  let fileQueue = [], activeStreamId = null, isCancelled = false;
  let timerInterval = null, timerStart = null;
  const uploadBtn = document.getElementById('uploadBtn');
  const stopBtn   = document.getElementById('stopBtn');
  const addBtn    = document.getElementById('addBtn');

  function setRunning(v) {
    uploadBtn.disabled = v;
    addBtn.disabled    = v;
    stopBtn.style.display = v ? 'inline-block' : 'none';
    stopBtn.disabled      = false;
    stopBtn.textContent   = '⛔ Stop';
    document.querySelectorAll('.remove-btn').forEach(b => b.disabled = v);
  }

  function startTimer() {
    timerStart = Date.now();
    document.getElementById('timer').style.display = 'block';
    timerInterval = setInterval(() => {
      document.getElementById('timerVal').textContent =
        ((Date.now() - timerStart) / 1000).toFixed(1) + 's';
    }, 100);
  }
  function stopTimer() { clearInterval(timerInterval); }

  function append(msg, cls) {
    const d  = document.getElementById('log');
    const el = document.createElement('div');
    if (cls) el.className = cls;
    el.textContent = msg;
    d.appendChild(el);
    d.parentElement.scrollTop = d.parentElement.scrollHeight;
  }

  function renderFileList() {
    const el = document.getElementById('file-list');
    if (!fileQueue.length) {
      el.innerHTML = '<div style="color:#555;font-size:.82rem;padding:.3rem 0">No files added.</div>';
      return;
    }
    el.innerHTML = fileQueue.map((f, i) => `
      <div class="file-row" id="file-row-${i}">
        <button class="remove-btn" onclick="removeFile(${i})">✕</button>
        <span class="file-name">${f.name}</span>
        <span class="file-size">${(f.size / 1024).toFixed(1)} KB</span>
        <span class="file-status" id="file-status-${i}">queued</span>
      </div>`).join('');
  }

  function addFiles(files) {
    for (const f of files) fileQueue.push(f);
    document.getElementById('hiddenInput').value = '';
    renderFileList();
  }
  function removeFile(i) { fileQueue.splice(i, 1); renderFileList(); }

  function setFileStatus(i, text, cls) {
    const el = document.getElementById('file-status-' + i);
    if (!el) return;
    el.textContent = text;
    el.className = 'file-status' + (cls ? ' ' + cls : '');
  }

  async function processFile(file, idx, provider) {
    setFileStatus(idx, 'uploading…', 'processing');
    const fd = new FormData();
    fd.append('file', file);
    fd.append('provider', provider);

    let res;
    try {
      res = await fetch('/api/reg-crawl/upload', { method: 'POST', body: fd });
    } catch (e) {
      append('[' + file.name + '] Upload failed: ' + e, 'error');
      setFileStatus(idx, '✗ failed', 'error');
      return false;
    }
    if (!res.ok) {
      append('[' + file.name + '] Server error', 'error');
      setFileStatus(idx, '✗ failed', 'error');
      return false;
    }
    const { stream_id } = await res.json();
    activeStreamId = stream_id;
    setFileStatus(idx, 'crawling…', 'processing');

    return new Promise(resolve => {
      const es = new EventSource('/api/reg-crawl/stream/' + stream_id);
      es.onmessage = e => {
        const msg = JSON.parse(e.data);
        if (msg.type === 'log') {
          append(msg.message, msg.is_error ? 'log-err' : 'log-line');
        } else if (msg.type === 'progress') {
          append('✓ [' + msg.processed + '/' + msg.total + '] ' + msg.message, '');
        } else if (msg.type === 'start') {
          append('[' + file.name + '] Starting — ' + msg.total + ' row(s) to crawl', '');
        } else if (msg.type === 'timer_start') {
          startTimer();
        } else if (msg.type === 'timer_stop') {
          stopTimer();
        } else if (msg.type === 'complete') {
          append(msg.message, 'done');
          setFileStatus(idx, '✓ done', 'done');
          window.location.href = '/api/reg-crawl/download/' + stream_id;
          es.close(); resolve(true);
        } else if (msg.type === 'cancelled') {
          append('⛔ Stopped — downloading partial results…', 'error');
          setFileStatus(idx, '⛔ stopped', 'error');
          window.location.href = '/api/reg-crawl/download/' + stream_id;
          es.close(); resolve(false);
        } else if (msg.type === 'error') {
          append('Error: ' + msg.message, 'error');
          setFileStatus(idx, '✗ error', 'error');
          es.close(); resolve(false);
        } else if (msg.type === 'stream_end') {
          es.close(); resolve(false);
        }
      };
      es.onerror = () => { es.close(); resolve(false); };
    });
  }

  async function startCrawl() {
    if (!fileQueue.length) { alert('Add at least one .xlsx file.'); return; }
    document.getElementById('log').innerHTML = '';
    isCancelled = false;
    setRunning(true);
    const provider = document.getElementById('provider').value;
    let ok = 0;
    for (let i = 0; i < fileQueue.length; i++) {
      if (isCancelled) { append('⛔ Queue stopped.', 'error'); break; }
      if (await processFile(fileQueue[i], i, provider)) ok++;
    }
    stopTimer();
    if (!isCancelled) append('Done — ' + ok + '/' + fileQueue.length + ' file(s) processed.', 'done');
    setRunning(false);
  }

  async function stopCrawl() {
    isCancelled = true;
    stopBtn.disabled  = true;
    stopBtn.textContent = '⛔ Stopping…';
    if (activeStreamId) {
      try { await fetch('/api/reg-crawl/cancel/' + activeStreamId, { method: 'POST' }); } catch(e) {}
    }
  }
</script>
</body>
</html>"""
    return html


@reg_crawl_bp.route("/api/reg-crawl/upload", methods=["POST"])
def reg_crawl_upload():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    if not f.filename.endswith(".xlsx"):
        return jsonify({"error": "Only .xlsx files are supported"}), 400

    provider   = request.form.get("provider", "gemini")
    file_bytes = f.read()
    filename   = f.filename
    stream_id  = uuid.uuid4().hex

    with _queues_lock:
        _queues[stream_id] = []

    threading.Thread(
        target=_run_reg_crawl,
        args=(stream_id, file_bytes, filename, provider),
        daemon=True,
    ).start()

    return jsonify({"stream_id": stream_id})


@reg_crawl_bp.route("/api/reg-crawl/stream/<stream_id>")
def reg_crawl_stream(stream_id):
    def generate():
        while True:
            with _queues_lock:
                queue = _queues.get(stream_id, [])
                events, _queues[stream_id] = queue[:], []
            for event in events:
                yield f"data: {json.dumps(event)}\n\n"
                if event.get("type") == "stream_end":
                    return
            time.sleep(0.3)
    return Response(
        generate(),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@reg_crawl_bp.route("/api/reg-crawl/download/<stream_id>")
def reg_crawl_download(stream_id):
    with _store_lock:
        info = _store.get(stream_id)
    if not info:
        return jsonify({"error": "File not ready or invalid stream_id"}), 404
    return send_file(
        info["path"],
        as_attachment=True,
        download_name=info["filename"],
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@reg_crawl_bp.route("/api/reg-crawl/cancel/<stream_id>", methods=["POST"])
def reg_crawl_cancel(stream_id):
    with _cancel_lock:
        evt = _cancel.get(stream_id)
    if evt:
        evt.set()
        return jsonify({"ok": True})
    return jsonify({"error": "No active job"}), 404
