"""
verify.py — Supplier verification layer

Reads an exported agentic Excel file, runs three web-search + LLM checks
per supplier row, and returns an annotated Excel with five new columns:
  Verification Notes | Company Exists | Supply Ties | Correct Component Supplied | URLs

Registered in app.py via:
    from verify import verify_bp
    app.register_blueprint(verify_bp)
"""

import io
import json
import os
import tempfile
import threading
import time
import uuid

import openpyxl
from flask import Blueprint, Response, jsonify, request, send_file
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from ai import call_ai, safe_parse_json
from scraper import web_search

# ── _export_files is imported from app at call-time to avoid circular imports ─

verify_bp = Blueprint("verify", __name__)

# ── Shared utilities ─────────────────────────────────────────────────────────

def _make_logger(log_fn):
    """Return a _log(msg, is_error) callable that prints and optionally pushes to SSE."""
    def _log(msg, is_error=False):
        print(msg)
        if log_fn:
            log_fn(msg, is_error)
    return _log


def _build_evidence(results: list[dict]) -> str:
    """Join scraped results into a single evidence block for LLM consumption."""
    if not results:
        return "No search results returned."
    return "\n---\n".join(f"SOURCE: {r['url']}\nCONTENT: {r['content']}" for r in results)


def _call_ai_json(prompt: str, provider: str, max_tokens: int,
                  required_key: str, log) -> dict | None:
    """Call the LLM with up to 3 attempts, returning the parsed dict if required_key is present."""
    for attempt in range(1, 4):
        try:
            raw = call_ai(prompt, provider, max_tokens=max_tokens)
            parsed = safe_parse_json(raw)
            if isinstance(parsed, dict) and required_key in parsed:
                return parsed
            log(f"[llm] Attempt {attempt}: response missing '{required_key}' — retrying", True)
        except Exception as e:
            log(f"[llm] Attempt {attempt} failed: {e}", True)
        if attempt < 3:
            time.sleep(5 * attempt)  # 5s after attempt 1, 10s after attempt 2
    log("[llm] All 3 attempts failed", True)
    return None


def _unique_urls(urls: list[str]) -> list[str]:
    """Deduplicate a flat list of URL strings while preserving order."""
    seen: set[str] = set()
    result = []
    for url in urls:
        if url not in seen:
            seen.add(url)
            result.append(url)
    return result


# Per-run queues: stream_id -> list of SSE event dicts
_verify_queues: dict = {}
_verify_queues_lock = threading.Lock()

# Snapshot temp file paths for mid-run partial downloads: stream_id -> file path
_snapshot_store: dict = {}
_snapshot_lock = threading.Lock()

# Cancel events: stream_id -> threading.Event (set to request cancellation)
_cancel_store: dict = {}
_cancel_lock = threading.Lock()


# ── LLM verification helper ───────────────────────────────────────────────────

def _ai_judge_all(evidence: str, company_name: str, supplies_to: str, components: str, provider: str, log_fn=None) -> dict:
    """
    Send combined evidence to the LLM once and ask all 3 verification questions together.
    Returns {"company_exists": bool, "supply_ties": bool, "correct_component": bool,
             "notes_exists": str, "notes_supply": str, "notes_component": str}.
    """
    _log = _make_logger(log_fn)

    prompt = f"""You are a supply chain fact-checker. Answer the three questions below using the web search evidence as your primary source. Only fall back to your training knowledge if the evidence is absent or inconclusive for a specific question.

EVIDENCE:
{evidence[:6000]}

Answer the following three questions about "{company_name}":

1. Does "{company_name}" exist as a real company that manufactures or supplies industrial/commercial products?
   - Accept common variations in naming: abbreviations, parent company names, subsidiary names, rebranded names, or regional name differences all count as the same company.

2. Is there a supply relationship between "{company_name}" and "{supplies_to}" related to "{components}"?
   - Accept direct supply contracts, but also count: documented collaborations, joint ventures, co-development agreements, or partnerships between the two companies that are relevant to the production or supply of "{components}".

3. Does "{company_name}" manufacture or produce "{components}"?

Return ONLY valid JSON (no markdown):
{{
  "company_exists": true or false,
  "source_exists": "web_evidence" or "training_knowledge",
  "supply_ties": true or false,
  "source_supply": "web_evidence" or "training_knowledge",
  "correct_component": true or false,
  "source_component": "web_evidence" or "training_knowledge",
  "notes_exists": "one sentence of reasoning citing the evidence used",
  "notes_supply": "one sentence of reasoning citing the evidence used",
  "notes_component": "one sentence of reasoning citing the evidence used"
}}

Set source_* to "training_knowledge" if the web evidence was absent or inconclusive and you fell back to your training knowledge. Otherwise set it to "web_evidence".
Only set a field to false if neither the web evidence nor your training knowledge supports it."""

    _log(f"[llm] Sending combined evidence to {provider} for '{company_name}' ({len(evidence.encode())} bytes)")
    parsed = _call_ai_json(prompt, provider, 500, "company_exists", _log)
    if parsed:
        _log(f"[llm] Response received for '{company_name}' — exists={parsed.get('company_exists')}, supply_ties={parsed.get('supply_ties')}, correct_component={parsed.get('correct_component')}")
        return {
            "company_exists":    bool(parsed.get("company_exists", False)),
            "supply_ties":       bool(parsed.get("supply_ties", False)),
            "correct_component": bool(parsed.get("correct_component", False)),
            "source_exists":     str(parsed.get("source_exists",    "training_knowledge")),
            "source_supply":     str(parsed.get("source_supply",    "training_knowledge")),
            "source_component":  str(parsed.get("source_component", "training_knowledge")),
            "notes_exists":      str(parsed.get("notes_exists", "")),
            "notes_supply":      str(parsed.get("notes_supply", "")),
            "notes_component":   str(parsed.get("notes_component", "")),
        }
    return {
        "company_exists": False, "supply_ties": False, "correct_component": False,
        "source_exists": "training_knowledge", "source_supply": "training_knowledge", "source_component": "training_knowledge",
        "notes_exists": "LLM call failed", "notes_supply": "LLM call failed", "notes_component": "LLM call failed",
    }



def verify_supplier_row(company_name: str, supplies_to: str, components: str, provider: str, log_fn=None) -> dict:
    """
    Run three web searches with a shared browser, combine all evidence, and send to
    the LLM in a single call covering all 3 verification questions.

    Returns:
        {
            "company_exists":     "Yes" | "No",
            "supply_ties":        "Yes" | "No",
            "correct_component":  "Yes" | "No",
            "notes":              str,
            "urls":               [str],
        }
    """
    _log = _make_logger(log_fn)
    component_query = components[:120] if components else "components"
    all_results: list[dict] = []
    verdict = {
        "company_exists": False, "supply_ties": False, "correct_component": False,
        "source_exists": "training_knowledge", "source_supply": "training_knowledge", "source_component": "training_knowledge",
        "notes_exists": "Browser error", "notes_supply": "Browser error", "notes_component": "Browser error",
    }

    # ── 3 broad searches via DDG + Crawl4AI ──────────────────────────────────
    r1 = web_search(f'"{company_name}" supplier manufacturer', n=10, max_scrape=4, log_fn=log_fn)
    r2 = web_search(f'"{company_name}" "{supplies_to}" supply partnership', n=10, max_scrape=4, log_fn=log_fn)
    r3 = web_search(f'"{company_name}" "{component_query}" manufacture supply', n=10, max_scrape=4, log_fn=log_fn)
    all_results = r1 + r2 + r3

    _log(f"[llm] Combined evidence: {len(all_results)} page(s) scraped across 3 queries")
    verdict = _ai_judge_all(_build_evidence(all_results), company_name, supplies_to, component_query, provider, log_fn=log_fn)

    unique_urls = _unique_urls([r["url"] for r in all_results])

    notes_parts = []
    if verdict["notes_exists"]:
        notes_parts.append(f"[Exists] {verdict['notes_exists']}")
    if verdict["notes_supply"]:
        notes_parts.append(f"[Supply] {verdict['notes_supply']}")
    if verdict["notes_component"]:
        notes_parts.append(f"[Component] {verdict['notes_component']}")

    return {
        "company_exists":    "Yes" if verdict["company_exists"] else "No",
        "supply_ties":       "Yes" if verdict["supply_ties"] else "No",
        "correct_component": "Yes" if verdict["correct_component"] else "No",
        "notes":             " | ".join(notes_parts),
        "notes_exists":      verdict["notes_exists"],
        "notes_supply":      verdict["notes_supply"],
        "notes_component":   verdict["notes_component"],
        "label_exists":      "Web Evidence" if verdict["source_exists"]    == "web_evidence" else "Training Data",
        "label_supply":      "Web Evidence" if verdict["source_supply"]    == "web_evidence" else "Training Data",
        "label_component":   "Web Evidence" if verdict["source_component"] == "web_evidence" else "Training Data",
        "urls":              unique_urls,
    }


# ── Workbook annotation ───────────────────────────────────────────────────────

VERIFY_HEADERS = [
    "Verification Notes",
    "Company Exists",
    "Supply Ties",
    "Correct Component Supplied",
    "URLs",
]

_HDR_FILL = PatternFill("solid", fgColor="1A237E")  # dark navy
_HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
_YES_FILL = PatternFill("solid", fgColor="C8E6C9")   # light green
_NO_FILL  = PatternFill("solid", fgColor="FFCDD2")   # light red


def _find_col(ws, header_name: str) -> int | None:
    """Return 1-based column index for a header in row 1, or None."""
    for cell in ws[1]:
        if cell.value and str(cell.value).strip().lower() == header_name.lower():
            return cell.column
    return None


_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

def _finalise_sheet(ws):
    """Auto-fit column widths, apply borders to all data cells, and enable AutoFilter."""
    # ── Column widths based on max content length per column ─────────────────
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                # Account for newlines in wrapped text: use longest line
                lines = str(cell.value).split("\n")
                max_len = max(max_len, max(len(line) for line in lines))
        # Add padding, cap at 80 chars so columns don't become unreadably wide
        ws.column_dimensions[col_letter].width = min(max_len + 2, 80)

    # ── Borders on all non-empty cells ────────────────────────────────────────
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                cell.border = _THIN_BORDER

    # ── Row bestFit so Excel auto-calculates height for wrapped text ──────────
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].bestFit = True

    # ── AutoFilter across all used columns ────────────────────────────────────
    ws.auto_filter.ref = ws.dimensions


def annotate_workbook(wb: openpyxl.Workbook, stream_id: str, provider: str) -> openpyxl.Workbook:
    """
    Iterate every tier sheet in the workbook, run verification per supplier row,
    and append the 5 verification columns. Pushes SSE progress events to the queue.
    """
    def push(event: dict):
        with _verify_queues_lock:
            if stream_id in _verify_queues:
                _verify_queues[stream_id].append(event)

    # Identify tier sheets (written by write_tier_sheet — names contain "tier")
    tier_sheets = [ws for ws in wb.worksheets if "tier" in ws.title.lower()]

    if not tier_sheets:
        push({"type": "status", "message": "⚠️ No tier sheets found in this workbook."})
        return wb

    # Count total supplier rows for progress reporting
    total_rows = 0
    for ws in tier_sheets:
        col_company = _find_col(ws, "Company Name")
        if col_company:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[col_company - 1]:
                    total_rows += 1

    push({"type": "start", "total": total_rows})
    push({"type": "timer_start"})
    processed = 0

    def log_fn(msg, is_error=False):
        push({"type": "log", "message": msg, "is_error": is_error})

    for ws in tier_sheets:
        col_company   = _find_col(ws, "Company Name")
        col_supplies  = _find_col(ws, "Supplies To")
        col_component = _find_col(ws, "Components Supplied")

        if not col_company:
            push({"type": "status", "message": f"⚠️ Skipping sheet '{ws.title}' — no 'Company Name' column"})
            continue

        # Find existing verification columns in the uploaded template
        col_ver_notes    = _find_col(ws, "Verification Notes")
        col_co_exists    = _find_col(ws, "Company Exists")
        col_supply_ties  = _find_col(ws, "Supply Ties Exist")
        col_correct_comp = _find_col(ws, "Correct Component Supplied")
        col_urls         = _find_col(ws, "URL")

        missing = [name for name, col in [
            ("Verification Notes", col_ver_notes),
            ("Company Exists",     col_co_exists),
            ("Supply Ties Exist",  col_supply_ties),
            ("Correct Component Supplied", col_correct_comp),
            ("URL",                col_urls),
        ] if col is None]
        if missing:
            push({"type": "log", "message": f"⚠️ Sheet '{ws.title}' — columns not found: {', '.join(missing)}", "is_error": True})

        # Set widths on existing verification columns
        for col, width in [
            (col_ver_notes,    45),
            (col_co_exists,    16),
            (col_supply_ties,  18),
            (col_correct_comp, 24),
            (col_urls,         50),
        ]:
            if col:
                ws.column_dimensions[get_column_letter(col)].width = width

        # Add 6 new columns after URL: label + source pairs for each of the 3 fields
        source_start_col = (col_urls + 1) if col_urls else (ws.max_column + 1)
        _SOURCE_COLS = [
            ("Company Exists Label",             "Source Company Exists"),
            ("Supply Ties Exists Label",          "Source Supply Ties Exists"),
            ("Correct Component Supplied Label",  "Source Correct Component Supplied"),
        ]
        for i, (label_hdr, source_hdr) in enumerate(_SOURCE_COLS):
            label_col  = source_start_col + i * 2
            source_col = source_start_col + i * 2 + 1
            for col_idx, hdr, width in [(label_col, label_hdr, 18), (source_col, source_hdr, 45)]:
                cell = ws.cell(row=1, column=col_idx, value=hdr)
                cell.fill = _HDR_FILL
                cell.font = _HDR_FONT
                ws.column_dimensions[cell.column_letter].width = width

        _WRAP = Alignment(wrap_text=True, vertical="top")

        # ── Phase 1: Collect all rows into a list (main thread, no workers yet) ──
        rows_to_process = []
        for row_idx in range(2, ws.max_row + 1):
            company_name = ws.cell(row=row_idx, column=col_company).value
            if not company_name:
                continue
            supplies_to = ws.cell(row=row_idx, column=col_supplies).value if col_supplies else ""
            components  = ws.cell(row=row_idx, column=col_component).value if col_component else ""
            rows_to_process.append((row_idx, str(company_name), str(supplies_to or ""), str(components or "")))

        # ── Phase 2: Load all rows onto the conveyor belt as futures ───────────
        from concurrent.futures import ThreadPoolExecutor, as_completed
        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = {
                executor.submit(verify_supplier_row, company, supplies_to, components, provider, log_fn): row_idx
                for row_idx, company, supplies_to, components in rows_to_process
            }

            # ── Phase 3: Write results back to sheet as workers finish ─────────
            for future in as_completed(futures):
                row_idx = futures[future]
                processed += 1
                try:
                    result = future.result()
                except Exception as e:
                    push({"type": "log", "message": f"[error] Row {row_idx} failed: {e}", "is_error": True})
                    continue

                push({
                    "type": "progress",
                    "row": processed,
                    "total": total_rows,
                    "message": f"[{processed}/{total_rows}] Row {row_idx} — Company Exists: {result['company_exists']} | Supply Ties: {result['supply_ties']} | Correct Component: {result['correct_component']}",
                })

                # Write to existing columns
                if col_ver_notes:
                    c = ws.cell(row=row_idx, column=col_ver_notes, value=result["notes"])
                    c.alignment = _WRAP
                if col_co_exists:
                    c = ws.cell(row=row_idx, column=col_co_exists, value=result["company_exists"])
                    c.alignment = _WRAP
                    c.fill = _YES_FILL if result["company_exists"] == "Yes" else _NO_FILL
                if col_supply_ties:
                    c = ws.cell(row=row_idx, column=col_supply_ties, value=result["supply_ties"])
                    c.alignment = _WRAP
                    c.fill = _YES_FILL if result["supply_ties"] == "Yes" else _NO_FILL
                if col_correct_comp:
                    c = ws.cell(row=row_idx, column=col_correct_comp, value=result["correct_component"])
                    c.alignment = _WRAP
                    c.fill = _YES_FILL if result["correct_component"] == "Yes" else _NO_FILL
                if col_urls:
                    c = ws.cell(row=row_idx, column=col_urls, value=", ".join(result["urls"]))
                    c.alignment = Alignment(wrap_text=False, vertical="top")

                # Write 6 new columns: label + source pairs
                for j, (notes_key, label_key) in enumerate([
                    ("notes_exists",    "label_exists"),
                    ("notes_supply",    "label_supply"),
                    ("notes_component", "label_component"),
                ]):
                    c = ws.cell(row=row_idx, column=source_start_col + j * 2, value=result[label_key])
                    c.alignment = _WRAP
                    c = ws.cell(row=row_idx, column=source_start_col + j * 2 + 1, value=result[notes_key])
                    c.alignment = _WRAP

                push({"type": "log", "message": f"[excel] Row {row_idx} written — exists={result['company_exists']}, supply={result['supply_ties']}, component={result['correct_component']}", "is_error": False})

                if processed % 10 == 0:
                    with _snapshot_lock:
                        snap_path = _snapshot_store.get(stream_id)
                    if snap_path:
                        try:
                            snap_buf = io.BytesIO()
                            wb.save(snap_buf)
                            with open(snap_path, "wb") as sf:
                                sf.write(snap_buf.getvalue())
                        except Exception:
                            pass

                # Check for user-requested cancellation
                with _cancel_lock:
                    cancel_evt = _cancel_store.get(stream_id)
                if cancel_evt and cancel_evt.is_set():
                    for f in futures:
                        f.cancel()
                    push({"type": "cancelled", "processed": processed})
                    break

    # ── Finalise all tier sheets: widths, borders, filter, row heights ───────
    for ws in tier_sheets:
        _finalise_sheet(ws)

    push({"type": "status", "message": f"✅ Verification complete — {processed} row(s) processed."})
    return wb


# ── Background worker ─────────────────────────────────────────────────────────

def _run_verification(stream_id: str, file_bytes: bytes, filename: str, provider: str):
    """Background thread: load workbook, annotate, save, signal file_ready."""
    from app import _export_files, _export_files_lock

    def push(event: dict):
        with _verify_queues_lock:
            if stream_id in _verify_queues:
                _verify_queues[stream_id].append(event)

    snap_path = None
    cancel_evt = threading.Event()
    try:
        with _cancel_lock:
            _cancel_store[stream_id] = cancel_evt

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_snap:
            snap_path = tmp_snap.name
        with _snapshot_lock:
            _snapshot_store[stream_id] = snap_path

        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        wb = annotate_workbook(wb, stream_id, provider)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        # Save to temp file for download
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(buf.read())
            tmp_path = tmp.name

        was_cancelled = cancel_evt.is_set()
        suffix = "_partial.xlsx" if was_cancelled else "_verified.xlsx"
        out_filename = filename.replace(".xlsx", suffix)
        file_id = uuid.uuid4().hex
        with _export_files_lock:
            _export_files[file_id] = {"path": tmp_path, "filename": out_filename}

        push({"type": "timer_stop"})
        push({"type": "file_ready", "file_id": file_id, "filename": out_filename})

    except Exception as e:
        push({"type": "error", "message": str(e)})
    finally:
        with _cancel_lock:
            _cancel_store.pop(stream_id, None)
        with _snapshot_lock:
            _snapshot_store.pop(stream_id, None)
        if snap_path:
            try:
                os.remove(snap_path)
            except OSError:
                pass
        push({"type": "done"})


# ── Routes ────────────────────────────────────────────────────────────────────

@verify_bp.route("/dev/verify")
def verify_page():
    """Upload page — black terminal UI matching the rest of the dev tools."""
    html = """<!DOCTYPE html>
<html>
<head>
  <title>Supplier Verification</title>
  <style>
    body { font-family: monospace; padding: 2rem; background: #111; color: #eee; }
    h2 span { color: #e8ff47; }
    input[type=file] { margin: 1rem 0; display: block; color: #eee; }
    button { background: #333; color: #eee; border: 1px solid #555; padding: .5rem 1.5rem; cursor: pointer; font-family: monospace; }
    button:hover { background: #444; }
    #stopBtn { background: #5c1a1a; border-color: #a33; display: none; margin-left: .5rem; }
    #stopBtn:hover { background: #7a2020; }
    #stopBtn:disabled { opacity: .5; cursor: not-allowed; }
    #log-container { height: 400px; overflow-y: auto; background: #1a1a1a; border: 1px solid #333; padding: .75rem; margin-top: 1.5rem; border-radius: 4px; }
    #log div { margin: 2px 0; font-size: .85rem; line-height: 1.6; }
    .done    { color: #4caf50; }
    .error   { color: #f44336; }
    .log-line { color: #888; font-size: .78rem; }
    .log-err  { color: #e57373; font-size: .78rem; }
    select { background: #222; color: #eee; border: 1px solid #555; padding: .3rem; font-family: monospace; margin-bottom: 1rem; }
    label  { font-size: .9rem; color: #aaa; }
    #timer { display: none; font-size: .85rem; color: #e8ff47; margin-top: .75rem; letter-spacing: .05em; }
    #timer span { font-weight: bold; }
    #snapshot-info { display: none; margin-top: .75rem; font-size: .82rem; color: #aaa; }
    #snapshot-info a { color: #47c8ff; text-decoration: none; }
    #snapshot-info a:hover { text-decoration: underline; }
  </style>
</head>
<body>
  <h2>Supplier <span>Verification</span></h2>
  <p>Upload an Excel file exported from the agentic pipeline. Each supplier row will be
     verified via web search across 3 dimensions.</p>

  <label>LLM Provider for judgement:</label><br>
  <select id="provider">
    <option value="gemini">Gemini (recommended — cheapest)</option>
    <option value="anthropic">Anthropic</option>
  </select>

  <input type="file" id="fileInput" accept=".xlsx">
  <button id="uploadBtn" onclick="startVerification()">Upload &amp; Verify</button>
  <button id="stopBtn" onclick="stopVerification()">⛔ Stop</button>

  <div id="timer">⏱ Elapsed: <span id="timerVal">0.0s</span></div>
  <div id="snapshot-info">💾 If connection drops: <a id="snapshot-link" href="#" target="_blank">download partial results</a></div>
  <div id="log-container"><div id="log"></div></div>

  <script>
    const log = document.getElementById('log');
    const logContainer = document.getElementById('log-container');
    const append = (msg, cls) => {
      const d = document.createElement('div');
      if (cls) d.className = cls;
      d.textContent = msg;
      log.appendChild(d);
      logContainer.scrollTop = logContainer.scrollHeight;
    };

    let timerInterval = null;
    let timerStart = null;
    let activeStreamId = null;

    const uploadBtn = document.getElementById('uploadBtn');
    const stopBtn   = document.getElementById('stopBtn');

    function setRunning(running) {
      uploadBtn.disabled = running;
      stopBtn.style.display = running ? 'inline-block' : 'none';
      stopBtn.disabled = false;
    }

    async function stopVerification() {
      if (!activeStreamId) return;
      stopBtn.disabled = true;
      stopBtn.textContent = '⛔ Stopping…';
      try {
        await fetch(`/api/verify/cancel/${activeStreamId}`, { method: 'POST' });
      } catch (e) {
        append('Could not reach cancel endpoint: ' + e, 'error');
      }
    }

    function startTimer() {
      timerStart = Date.now();
      const disp = document.getElementById('timer');
      const val  = document.getElementById('timerVal');
      disp.style.display = 'block';
      val.textContent = '0.0s';
      if (timerInterval) clearInterval(timerInterval);
      timerInterval = setInterval(() => {
        val.textContent = ((Date.now() - timerStart) / 1000).toFixed(1) + 's';
      }, 100);
    }

    function stopTimer() {
      if (timerInterval) { clearInterval(timerInterval); timerInterval = null; }
      if (!timerStart) return;
      const elapsed = ((Date.now() - timerStart) / 1000).toFixed(1);
      document.getElementById('timerVal').textContent = elapsed + 's ✓';
    }

    async function startVerification() {
      const file = document.getElementById('fileInput').files[0];
      if (!file) { append('⚠️ Please select an Excel file first.'); return; }

      const provider = document.getElementById('provider').value;
      append(`Uploading ${file.name}…`);
      setRunning(true);

      const fd = new FormData();
      fd.append('file', file);
      fd.append('provider', provider);

      let streamId;
      try {
        const res = await fetch('/dev/verify/run', { method: 'POST', body: fd });
        const data = await res.json();
        if (data.error) { append('Error: ' + data.error, 'error'); setRunning(false); return; }
        streamId = data.stream_id;
        activeStreamId = streamId;
      } catch (e) {
        append('Upload failed: ' + e, 'error'); setRunning(false); return;
      }

      append('Upload complete — starting verification…');
      const snapshotLink = document.getElementById('snapshot-link');
      snapshotLink.href = `/dev/verify/snapshot/${streamId}`;
      document.getElementById('snapshot-info').style.display = 'block';

      const evtSource = new EventSource(`/dev/verify/stream/${streamId}`);

      evtSource.onmessage = (e) => {
        const msg = JSON.parse(e.data);
        if (msg.type === 'timer_start') {
          startTimer();
        } else if (msg.type === 'timer_stop') {
          stopTimer();
        } else if (msg.type === 'start') {
          append(`Processing ${msg.total} supplier row(s)…`);
        } else if (msg.type === 'progress') {
          append(msg.message);
        } else if (msg.type === 'status') {
          append(msg.message);
        } else if (msg.type === 'log') {
          append(msg.message, msg.is_error ? 'log-err' : 'log-line');
        } else if (msg.type === 'cancelled') {
          stopTimer();
          append(`⛔ Stopped after ${msg.processed} row(s) — downloading partial results…`, 'error');
          setRunning(false);
          stopBtn.textContent = '⛔ Stop';
        } else if (msg.type === 'file_ready') {
          append(`✓ Done — downloading ${msg.filename}`, 'done');
          const a = document.createElement('a');
          a.href = `/dev/verify/download/${msg.file_id}`;
          a.download = msg.filename;
          document.body.appendChild(a);
          a.click();
          document.body.removeChild(a);
          setRunning(false);
          evtSource.close();
        } else if (msg.type === 'error') {
          stopTimer();
          append('Error: ' + msg.message, 'error');
          setRunning(false);
          evtSource.close();
        } else if (msg.type === 'done') {
          setRunning(false);
          evtSource.close();
        }
      };

      evtSource.onerror = () => {
        stopTimer();
        append('Connection lost.', 'error');
        setRunning(false);
        evtSource.close();
      };
    }
  </script>
</body>
</html>"""
    return Response(html, mimetype="text/html")


@verify_bp.route("/api/verify/cancel/<stream_id>", methods=["POST"])
def verify_cancel(stream_id):
    """Signal the running verification job to stop after the current row."""
    with _cancel_lock:
        evt = _cancel_store.get(stream_id)
    if evt:
        evt.set()
        return jsonify({"ok": True})
    return jsonify({"error": "No active job found for this stream_id"}), 404


@verify_bp.route("/dev/verify/snapshot/<stream_id>")
def verify_snapshot(stream_id):
    """Download current partial results for a running verification job."""
    with _snapshot_lock:
        snap_path = _snapshot_store.get(stream_id)

    if not snap_path or not os.path.exists(snap_path) or os.path.getsize(snap_path) == 0:
        return jsonify({"error": "No snapshot available yet — job may not have processed 10 rows yet, or has already completed"}), 404

    buf = io.BytesIO()
    with open(snap_path, "rb") as f:
        buf.write(f.read())
    buf.seek(0)

    return send_file(
        buf,
        as_attachment=True,
        download_name=f"partial_verify_{stream_id[:8]}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@verify_bp.route("/dev/verify/run", methods=["POST"])
def verify_run():
    """Accept Excel upload, start background verification, return stream_id."""
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    f = request.files["file"]
    if not f.filename.endswith(".xlsx"):
        return jsonify({"error": "Only .xlsx files are supported"}), 400

    provider = request.form.get("provider", "gemini")
    file_bytes = f.read()
    filename = f.filename
    stream_id = uuid.uuid4().hex

    with _verify_queues_lock:
        _verify_queues[stream_id] = []

    thread = threading.Thread(
        target=_run_verification,
        args=(stream_id, file_bytes, filename, provider),
        daemon=True,
    )
    thread.start()

    return jsonify({"stream_id": stream_id})


@verify_bp.route("/dev/verify/stream/<stream_id>")
def verify_stream(stream_id):
    """SSE stream for a running verification job."""
    import time as _time

    def generate():
        while True:
            with _verify_queues_lock:
                queue = _verify_queues.get(stream_id, [])
                events, _verify_queues[stream_id] = queue[:], []

            for event in events:
                yield f"data: {json.dumps(event)}\n\n"
                if event.get("type") == "done":
                    with _verify_queues_lock:
                        _verify_queues.pop(stream_id, None)
                    return

            if not events:
                yield ": keepalive\n\n"
            _time.sleep(0.5)

    return Response(generate(), mimetype="text/event-stream", headers={
        "Cache-Control": "no-cache",
        "X-Accel-Buffering": "no",
    })


@verify_bp.route("/dev/verify/download/<file_id>")
def verify_download(file_id):
    """Download the annotated Excel file."""
    from app import _export_files, _export_files_lock

    with _export_files_lock:
        entry = _export_files.pop(file_id, None)

    if not entry:
        return jsonify({"error": "File not found or already downloaded"}), 404

    path     = entry["path"]
    filename = entry["filename"]

    response = send_file(
        path,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )

    @response.call_on_close
    def cleanup():
        try:
            os.remove(path)
        except OSError:
            pass

    return response
